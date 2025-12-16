#!/usr/bin/env python3
"""Export OPFL Excel scores to JSON for web display."""

import json
import re
import os
import sys
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import nflreadpy as nfl
import openpyxl

# Add project root to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from opfl import OPFLScorer, parse_roster_from_excel

OWNER_TO_CODE = {
    'KIRK/DAVID': 'K/D',
    'STEVE L.': 'STL',
    'JOHN': 'JOH',
    'KEVIN': 'KEV',
    'DANNY/JOEY': 'D/J',
    'GREG/GRIFFIN': 'G/G',
    'ERIC/JEFF': 'E/J',
    'ANDREW': 'AND',
    'WES/BILL': 'W/B',
    'KEMP/A/M': 'KAM',
    'JARRETT/MATT': 'J/M',
    'ADAM': 'ADA',
}

CODE_TO_OWNER = {v: k for k, v in OWNER_TO_CODE.items()}
ALL_TEAMS = list(OWNER_TO_CODE.values())
TEAM_COLUMNS = [4, 7, 10, 13, 16, 19]
POSITIONS = ['QB', 'RB', 'WR', 'TE', 'K', 'DF', 'HC']
TRADE_DEADLINE_WEEK = 12
REGULAR_SEASON_WEEKS = 15
PLAYOFF_WEEKS = [16, 17]

# 2025 OPFL Schedule - Team numbers mapped to abbreviations
# (1) Kemp/A/M, (2) Danny/Joey, (3) Steve L., (4) Kirk/David, (5) Kevin, (6) Eric/Jeff
# (7) Wes/Bill, (8) Andrew, (9) Jarrett/Matt, (10) Adam, (11) John, (12) Greg/Griffin
TEAM_NUMBER_MAP = {
    1: 'KAM', 2: 'D/J', 3: 'STL', 4: 'K/D', 5: 'KEV', 6: 'E/J',
    7: 'W/B', 8: 'AND', 9: 'J/M', 10: 'ADA', 11: 'JOH', 12: 'G/G'
}

# Schedule: dict of week -> list of matchups (team1_num, team2_num)
SCHEDULE = {
    1:  [[1,2], [3,4], [5,6], [7,8], [9,10], [11,12]],
    2:  [[1,3], [2,4], [5,7], [6,8], [9,11], [10,12]],
    3:  [[1,4], [2,5], [3,6], [7,10], [8,11], [9,12]],
    4:  [[1,5], [2,6], [3,7], [4,9], [8,12], [10,11]],
    5:  [[1,6], [2,7], [3,12], [4,10], [5,11], [8,9]],
    6:  [[1,7], [2,8], [3,9], [4,12], [5,10], [6,11]],
    7:  [[1,8], [2,9], [3,10], [4,11], [5,12], [6,7]],
    8:  [[1,9], [2,10], [3,11], [4,6], [5,8], [7,12]],
    9:  [[1,10], [2,11], [3,8], [4,5], [7,9], [6,12]],
    10: [[1,11], [2,12], [3,5], [4,7], [6,9], [8,10]],
    11: [[1,12], [2,3], [4,8], [5,9], [6,10], [7,11]],
    12: [[1,2], [3,4], [5,6], [7,8], [9,10], [11,12]],
    13: [[1,3], [2,4], [5,7], [6,8], [9,11], [10,12]],
    14: [[1,4], [2,5], [3,6], [7,10], [8,11], [9,12]],
    15: [[1,5], [2,6], [3,7], [4,9], [8,12], [10,11]]
}


def parse_player_name(cell_value):
    if not cell_value:
        return "", ""
    cell_value = str(cell_value).strip()
    match = re.match(r'^(.+?)\s*\(([A-Za-z]{2,3})\)$', cell_value)
    if match:
        return match.group(1).strip(), match.group(2).upper()
    return cell_value, ""


def is_valid_player_name(name, position):
    """Check if a player name is valid (not just a number or empty)."""
    if not name:
        return False
    
    # Filter out numeric entries (scores, jersey numbers, etc.)
    # These can appear as integers, floats, or string representations thereof
    try:
        float(name)
        # If we get here, the name is purely numeric - not a valid player/coach name
        return False
    except (ValueError, TypeError):
        pass
    
    # Filter out phone numbers (patterns like "325-1289" or "325-1289 (J)")
    import re
    if re.match(r'^\d{3}-\d{4}', name):
        return False
    
    # For HC position, do additional validation
    if position == 'HC':
        # Head coach names should contain letters
        if not any(c.isalpha() for c in name):
            return False
        # Name should start with a letter (coach names don't start with numbers)
        if not name[0].isalpha():
            return False
    
    return True


def find_position_rows(ws, start_row, end_row):
    position_rows = {}
    current_position = None
    for row in range(start_row, min(end_row + 1, ws.max_row + 1)):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value:
            cell_str = str(cell_value).strip().upper()
            if cell_str in POSITIONS:
                current_position = cell_str
                if current_position not in position_rows:
                    position_rows[current_position] = []
                position_rows[current_position].append(row)
        elif current_position:
            has_content = any(ws.cell(row=row, column=c).value for c in TEAM_COLUMNS)
            if has_content:
                position_rows[current_position].append(row)
    return position_rows


def extract_team_name(header_value):
    if not header_value:
        return "", ""
    match = re.match(r'^(.+?)\s*\((\d+)\)$', str(header_value).strip())
    if match:
        return match.group(1).strip(), match.group(2)
    return str(header_value).strip(), ""


def export_week(ws, week_num, excel_path, season=2025):
    """Export a week's data, using autoscorer to calculate ALL player scores."""
    teams_data = []
    
    # Use autoscorer to get scores for ALL players (starters and bench)
    sheet_name = f"W{week_num}"
    try:
        fantasy_teams = parse_roster_from_excel(excel_path, sheet_name)
        scorer = OPFLScorer(season, week_num)
        
        for team in fantasy_teams:
            team_name = team.name
            abbrev = OWNER_TO_CODE.get(team_name, team_name[:3].upper())
            roster = []
            total_score = 0.0
            
            # Score ALL players on the team
            scores = scorer.score_fantasy_team(team, starters_only=False)
            
            for position, player_scores in scores.items():
                for ps in player_scores:
                    if is_valid_player_name(ps.name, position):
                        roster.append({
                            'name': ps.name,
                            'nfl_team': ps.team,
                            'position': position,
                            'score': round(ps.total_points, 1),
                            'starter': ps.is_starter,
                        })
                        # Only starters count toward team total
                        if ps.is_starter:
                            total_score += ps.total_points
            
            if roster:
                teams_data.append({
                    'name': team_name,
                    'owner': team_name,
                    'abbrev': abbrev,
                    'roster': roster,
                    'total_score': round(total_score, 1),
                })
        
        print(f"  Week {week_num}: Scored {len(teams_data)} teams with autoscorer")
        
    except Exception as e:
        print(f"  Week {week_num}: Autoscorer failed ({e}), falling back to Excel scores")
        # Fallback to reading from Excel if autoscorer fails
        teams_data = export_week_from_excel(ws, week_num)
    
    sorted_by_score = sorted(teams_data, key=lambda t: t['total_score'], reverse=True)
    for rank, team in enumerate(sorted_by_score, 1):
        team['score_rank'] = rank
    
    return {
        'week': week_num,
        'teams': teams_data,
        'has_scores': any(t['total_score'] > 0 for t in teams_data),
    }


def export_week_from_excel(ws, week_num):
    """Fallback: Export week data from Excel only (no backup scores)."""
    teams_data = []
    blocks = [(1, 1, 38), (39, 39, 80)]
    
    for header_row, start_row, end_row in blocks:
        position_rows = find_position_rows(ws, start_row, end_row)
        
        for col in TEAM_COLUMNS:
            header = ws.cell(row=header_row, column=col).value
            if not header:
                continue
            
            team_name, standing = extract_team_name(header)
            abbrev = OWNER_TO_CODE.get(team_name, team_name[:3].upper())
            star_col = col - 1
            roster = []
            total_score = 0.0
            
            for position, rows in position_rows.items():
                for row in rows:
                    player_cell = ws.cell(row=row, column=col)
                    score_cell = ws.cell(row=row, column=col - 2)
                    star_cell = ws.cell(row=row, column=star_col)
                    
                    if player_cell.value:
                        player_name, nfl_team = parse_player_name(str(player_cell.value))
                        is_starter = star_cell.value == '*'
                        score = float(score_cell.value) if score_cell.value else 0.0
                        
                        if player_name and is_valid_player_name(player_name, position):
                            roster.append({
                                'name': player_name,
                                'nfl_team': nfl_team,
                                'position': position,
                                'score': score,
                                'starter': is_starter,
                            })
                            if is_starter:
                                total_score += score
            
            if roster:
                teams_data.append({
                    'name': team_name,
                    'owner': team_name,
                    'abbrev': abbrev,
                    'roster': roster,
                    'total_score': round(total_score, 1),
                })
    
    return teams_data


def get_current_nfl_week():
    return nfl.get_current_week()


def get_game_times(season=2025):
    try:
        schedule = nfl.load_schedules(seasons=season)
        game_times = {}
        for week in range(1, 19):
            week_games = schedule.filter(schedule['week'] == week)
            if week_games.height == 0:
                continue
            game_times[week] = {}
            for row in week_games.iter_rows(named=True):
                game_date = row.get('gameday', '')
                game_time = row.get('gametime', '')
                if game_date and game_time:
                    try:
                        dt = datetime.strptime(f"{game_date} {game_time}", "%Y-%m-%d %H:%M")
                        kickoff_iso = dt.strftime("%Y-%m-%dT%H:%M:00-05:00")
                        if row.get('home_team'):
                            game_times[week][row['home_team']] = kickoff_iso
                        if row.get('away_team'):
                            game_times[week][row['away_team']] = kickoff_iso
                    except (ValueError, TypeError):
                        pass
        return game_times
    except Exception as e:
        print(f"Warning: Could not load game times: {e}")
        return {}


def parse_draft_picks(excel_path):
    DEFAULT_PRESEASON = list(range(1, 7))
    DEFAULT_WAIVER = list(range(1, 4))
    SEASONS = ['2026', '2027', '2028']
    
    picks = {}
    for team in ALL_TEAMS:
        picks[team] = {}
        for season in SEASONS:
            picks[team][season] = {
                'preseason': [(r, team) for r in DEFAULT_PRESEASON],
                'waiver': [(r, team) for r in DEFAULT_WAIVER],
            }
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['Future Traded Picks']
    except Exception as e:
        print(f"Warning: Could not load traded picks: {e}")
        return format_picks_for_output(picks)
    
    pattern = r"([A-Za-z/\.\s]+)\s+holds?\s+([A-Za-z/\.\s]+)'s\s+(\d{4})\s+#(\d+)\s+pick"
    
    for row in range(3, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=2).value
        if not cell_value:
            continue
        match = re.search(pattern, str(cell_value).strip(), re.IGNORECASE)
        if match:
            holder_name = match.group(1).strip()
            original_name = match.group(2).strip()
            season = match.group(3)
            round_num = int(match.group(4))
            
            holder = OWNER_TO_CODE.get(holder_name)
            original = OWNER_TO_CODE.get(original_name)
            
            if not holder or not original:
                for name, code in OWNER_TO_CODE.items():
                    if holder_name.upper() in name.upper() or name.upper() in holder_name.upper():
                        holder = code
                    if original_name.upper() in name.upper() or name.upper() in original_name.upper():
                        original = code
            
            if holder and original and season in SEASONS:
                draft_type = 'preseason' if round_num <= 6 else 'waiver'
                original_picks = picks[original][season][draft_type]
                for i, (r, owner) in enumerate(original_picks):
                    if r == round_num and owner == original:
                        original_picks.pop(i)
                        break
                picks[holder][season][draft_type].append((round_num, original))
    
    wb.close()
    return format_picks_for_output(picks)


def format_picks_for_output(picks):
    formatted = {}
    for team in ALL_TEAMS:
        formatted[team] = {}
        for season in picks[team]:
            formatted[team][season] = {}
            for draft_type in picks[team][season]:
                team_picks = sorted(picks[team][season][draft_type], key=lambda x: (x[0], x[1]))
                formatted[team][season][draft_type] = [
                    {'round': r, 'from': owner, 'own': owner == team}
                    for r, owner in team_picks
                ]
    return formatted


def get_existing_banners(banners_dir):
    """Get list of banner images from directory, sorted by year descending."""
    if not os.path.exists(banners_dir):
        return []
    images = [f for f in os.listdir(banners_dir) if f.endswith('.png')]
    def get_year(filename):
        try:
            return int(filename.replace('.png', ''))
        except ValueError:
            return 0
    return sorted(images, key=get_year, reverse=True)


def export_all_weeks(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    weeks = []
    standings = {}
    
    week_sheets = []
    for sheet_name in wb.sheetnames:
        match = re.match(r'^W(\d+)$', sheet_name)
        if match:
            week_sheets.append((int(match.group(1)), sheet_name))
    week_sheets.sort(key=lambda x: x[0])
    
    print("Scoring all weeks with autoscorer (this may take a moment)...")
    for week_num, sheet_name in week_sheets:
        weeks.append(export_week(wb[sheet_name], week_num, excel_path))
    
    current_nfl_week = get_current_nfl_week()
    print(f"Current NFL week: {current_nfl_week}, standings include weeks 1-{current_nfl_week - 1}")
    
    for week_data in weeks:
        if not week_data.get('has_scores', False) or week_data['week'] >= current_nfl_week:
            continue
        
        week_num = week_data['week']
        
        # Build team score lookup for this week
        team_scores = {}
        for team in week_data['teams']:
            abbrev = team['abbrev']
            team_scores[abbrev] = team['total_score']
            if abbrev not in standings:
                standings[abbrev] = {
                    'name': team['name'],
                    'owner': team['owner'],
                    'abbrev': abbrev,
                    'rank_points': 0.0,
                    'wins': 0,
                    'losses': 0,
                    'ties': 0,
                    'top_half': 0,
                    'points_for': 0.0,
                    'points_against': 0.0,
                }
            standings[abbrev]['points_for'] += team['total_score']
        
        # Calculate wins/losses/ties based on schedule matchups
        if week_num in SCHEDULE:
            for team1_num, team2_num in SCHEDULE[week_num]:
                abbrev1 = TEAM_NUMBER_MAP.get(team1_num)
                abbrev2 = TEAM_NUMBER_MAP.get(team2_num)
                
                if abbrev1 in team_scores and abbrev2 in team_scores:
                    score1 = team_scores[abbrev1]
                    score2 = team_scores[abbrev2]
                    
                    # Update points against
                    if abbrev1 in standings:
                        standings[abbrev1]['points_against'] += score2
                    if abbrev2 in standings:
                        standings[abbrev2]['points_against'] += score1
                    
                    # Determine win/loss/tie
                    if score1 > score2:
                        if abbrev1 in standings:
                            standings[abbrev1]['wins'] += 1
                            standings[abbrev1]['rank_points'] += 1
                        if abbrev2 in standings:
                            standings[abbrev2]['losses'] += 1
                    elif score2 > score1:
                        if abbrev2 in standings:
                            standings[abbrev2]['wins'] += 1
                            standings[abbrev2]['rank_points'] += 1
                        if abbrev1 in standings:
                            standings[abbrev1]['losses'] += 1
                    else:
                        # Tie
                        if abbrev1 in standings:
                            standings[abbrev1]['ties'] += 1
                            standings[abbrev1]['rank_points'] += 0.5
                        if abbrev2 in standings:
                            standings[abbrev2]['ties'] += 1
                            standings[abbrev2]['rank_points'] += 0.5
        
        # Calculate top 6 scoring bonus (0.5 points each)
        teams_by_score = sorted(week_data['teams'], key=lambda x: x['total_score'], reverse=True)
        current_rank = 1
        i = 0
        while i < len(teams_by_score):
            current_score = teams_by_score[i]['total_score']
            tied_teams = []
            while i < len(teams_by_score) and teams_by_score[i]['total_score'] == current_score:
                tied_teams.append(teams_by_score[i])
                i += 1
            
            tied_positions = list(range(current_rank, current_rank + len(tied_teams)))
            positions_in_top6 = [p for p in tied_positions if p <= 6]
            
            if positions_in_top6:
                points_per_team = (0.5 * len(positions_in_top6)) / len(tied_teams)
                for team in tied_teams:
                    standings[team['abbrev']]['rank_points'] += points_per_team
                    standings[team['abbrev']]['top_half'] += 1
            
            current_rank += len(tied_teams)
    
    wb.close()
    
    current_nfl_week = get_current_nfl_week()
    sorted_standings = sorted(standings.values(), key=lambda x: (x['rank_points'], x['points_for']), reverse=True)
    
    # Compute playoff data if regular season is complete
    playoffs = None
    if current_nfl_week > REGULAR_SEASON_WEEKS and len(sorted_standings) >= 4:
        playoffs = compute_playoff_data(weeks, sorted_standings, current_nfl_week)
    
    return {
        'updated_at': datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z'),
        'season': 2025,
        'current_week': current_nfl_week,
        'regular_season_weeks': REGULAR_SEASON_WEEKS,
        'weeks': weeks,
        'standings': sorted_standings,
        'playoffs': playoffs,
        'game_times': get_game_times(2025),
        'trade_deadline_week': TRADE_DEADLINE_WEEK,
        'taxi_squads': {team: [] for team in ALL_TEAMS},
    }


def load_pending_trades():
    path = Path(__file__).parent.parent / 'data' / 'pending_trades.json'
    if path.exists():
        with open(path) as f:
            return json.load(f).get('trades', [])
    return []


def compute_playoff_data(weeks_data, standings, current_nfl_week):
    """
    Compute playoff bracket and Jamboree data.
    
    Playoffs:
    - Week 16: Semifinals (1 vs 4, 2 vs 3)
    - Week 17: Oakland Bowl (winners), 3rd place game (losers)
    
    Jamboree:
    - All non-playoff teams (seeds 5-12)
    - Winner is whoever scores most total points over weeks 16-17
    """
    # Get final regular season standings (after week 15)
    if not standings or len(standings) < 4:
        return None
    
    # Playoff teams are top 4 seeds
    playoff_teams = [s['abbrev'] for s in standings[:4]]
    jamboree_teams = [s['abbrev'] for s in standings[4:]]
    
    # Create week lookup
    week_data_map = {w['week']: w for w in weeks_data}
    
    # Initialize playoff structure
    playoffs = {
        'playoff_teams': playoff_teams,
        'jamboree_teams': jamboree_teams,
        'seeds': {standings[i]['abbrev']: i + 1 for i in range(len(standings))},
        'week_16': {
            'semifinal_1': {  # 1 vs 4
                'higher_seed': standings[0]['abbrev'],
                'lower_seed': standings[3]['abbrev'],
                'higher_score': 0,
                'lower_score': 0,
                'winner': None,
                'loser': None,
            },
            'semifinal_2': {  # 2 vs 3
                'higher_seed': standings[1]['abbrev'],
                'lower_seed': standings[2]['abbrev'],
                'higher_score': 0,
                'lower_score': 0,
                'winner': None,
                'loser': None,
            },
        },
        'week_17': {
            'championship': {  # Oakland Bowl
                'team1': None,
                'team2': None,
                'score1': 0,
                'score2': 0,
                'winner': None,
                'loser': None,
            },
            'third_place': {
                'team1': None,
                'team2': None,
                'score1': 0,
                'score2': 0,
                'winner': None,
                'loser': None,
            },
        },
        'jamboree': {
            'standings': [],  # List of {abbrev, name, week_16_score, week_17_score, total}
            'winner': None,
        },
    }
    
    # Get week 16 scores
    week_16_data = week_data_map.get(16)
    if week_16_data and week_16_data.get('teams'):
        team_scores_16 = {t['abbrev']: t['total_score'] for t in week_16_data['teams']}
        
        # Update semifinal 1 (1 vs 4)
        sf1 = playoffs['week_16']['semifinal_1']
        sf1['higher_score'] = team_scores_16.get(sf1['higher_seed'], 0)
        sf1['lower_score'] = team_scores_16.get(sf1['lower_seed'], 0)
        if sf1['higher_score'] > 0 or sf1['lower_score'] > 0:
            if sf1['higher_score'] >= sf1['lower_score']:
                sf1['winner'] = sf1['higher_seed']
                sf1['loser'] = sf1['lower_seed']
            else:
                sf1['winner'] = sf1['lower_seed']
                sf1['loser'] = sf1['higher_seed']
        
        # Update semifinal 2 (2 vs 3)
        sf2 = playoffs['week_16']['semifinal_2']
        sf2['higher_score'] = team_scores_16.get(sf2['higher_seed'], 0)
        sf2['lower_score'] = team_scores_16.get(sf2['lower_seed'], 0)
        if sf2['higher_score'] > 0 or sf2['lower_score'] > 0:
            if sf2['higher_score'] >= sf2['lower_score']:
                sf2['winner'] = sf2['higher_seed']
                sf2['loser'] = sf2['lower_seed']
            else:
                sf2['winner'] = sf2['lower_seed']
                sf2['loser'] = sf2['higher_seed']
        
        # Set up week 17 matchups based on week 16 results
        if sf1['winner'] and sf2['winner']:
            playoffs['week_17']['championship']['team1'] = sf1['winner']
            playoffs['week_17']['championship']['team2'] = sf2['winner']
            playoffs['week_17']['third_place']['team1'] = sf1['loser']
            playoffs['week_17']['third_place']['team2'] = sf2['loser']
    
    # Get week 17 scores
    week_17_data = week_data_map.get(17)
    if week_17_data and week_17_data.get('teams'):
        team_scores_17 = {t['abbrev']: t['total_score'] for t in week_17_data['teams']}
        
        # Update championship (Oakland Bowl)
        champ = playoffs['week_17']['championship']
        if champ['team1'] and champ['team2']:
            champ['score1'] = team_scores_17.get(champ['team1'], 0)
            champ['score2'] = team_scores_17.get(champ['team2'], 0)
            if champ['score1'] > 0 or champ['score2'] > 0:
                if champ['score1'] >= champ['score2']:
                    champ['winner'] = champ['team1']
                    champ['loser'] = champ['team2']
                else:
                    champ['winner'] = champ['team2']
                    champ['loser'] = champ['team1']
        
        # Update third place game
        third = playoffs['week_17']['third_place']
        if third['team1'] and third['team2']:
            third['score1'] = team_scores_17.get(third['team1'], 0)
            third['score2'] = team_scores_17.get(third['team2'], 0)
            if third['score1'] > 0 or third['score2'] > 0:
                if third['score1'] >= third['score2']:
                    third['winner'] = third['team1']
                    third['loser'] = third['team2']
                else:
                    third['winner'] = third['team2']
                    third['loser'] = third['team1']
    
    # Compute Jamboree standings (combined scores for weeks 16-17)
    jamboree_standings = []
    for abbrev in jamboree_teams:
        team_info = next((s for s in standings if s['abbrev'] == abbrev), None)
        if not team_info:
            continue
        
        week_16_score = 0
        week_17_score = 0
        
        if week_16_data and week_16_data.get('teams'):
            team_16 = next((t for t in week_16_data['teams'] if t['abbrev'] == abbrev), None)
            if team_16:
                week_16_score = team_16.get('total_score', 0)
        
        if week_17_data and week_17_data.get('teams'):
            team_17 = next((t for t in week_17_data['teams'] if t['abbrev'] == abbrev), None)
            if team_17:
                week_17_score = team_17.get('total_score', 0)
        
        jamboree_standings.append({
            'abbrev': abbrev,
            'name': team_info['name'],
            'week_16_score': round(week_16_score, 1),
            'week_17_score': round(week_17_score, 1),
            'total': round(week_16_score + week_17_score, 1),
        })
    
    # Sort by total score descending
    jamboree_standings.sort(key=lambda x: x['total'], reverse=True)
    playoffs['jamboree']['standings'] = jamboree_standings
    
    # Determine Jamboree winner if week 17 is complete
    if jamboree_standings and current_nfl_week > 17:
        playoffs['jamboree']['winner'] = jamboree_standings[0]['abbrev']
    
    return playoffs


def main():
    project_dir = Path(__file__).parent.parent
    
    excel_path = None
    for fname in ['Griff OPFL Scoring 2025.xlsx', 'OPFL Scoring 2025.xlsx']:
        if (project_dir / fname).exists():
            excel_path = project_dir / fname
            break
    
    if not excel_path:
        print('Error: No OPFL Excel file found!')
        return
    
    output_path = project_dir / 'web' / 'data.json'
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    print(f'Exporting {excel_path} to {output_path}...')
    
    data = export_all_weeks(str(excel_path))
    data['pending_trades'] = load_pending_trades()
    
    draft_picks_path = project_dir / 'OPFL Draft & Future Traded Picks.xlsx'
    if draft_picks_path.exists():
        print('Parsing draft picks...')
        data['draft_picks'] = parse_draft_picks(str(draft_picks_path))
    
    banners_dir = project_dir / 'web' / 'images' / 'banners'
    if banners_dir.exists():
        print('Loading banners...')
        data['banners'] = get_existing_banners(str(banners_dir))
    
    with open(output_path, 'w') as f:
        json.dump(data, f, indent=2)
    
    print(f'Exported {len(data["weeks"])} weeks')
    print(f'Standings: {len(data["standings"])} teams')
    if 'banners' in data:
        print(f'Banners: {len(data["banners"])} images')
    print(f'Updated at: {data["updated_at"]}')


if __name__ == '__main__':
    main()
