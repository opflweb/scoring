#!/usr/bin/env python3
"""Migrate OPFL Excel data to JSON format."""

import json
import re
from pathlib import Path
import openpyxl

# OPFL positions
POSITIONS = ['QB', 'RB', 'WR', 'TE', 'K', 'DF', 'HC']

# Team columns in OPFL Excel (D, G, J, M, P, S)
TEAM_COLUMNS = [4, 7, 10, 13, 16, 19]

# Owner name to team code mapping
OWNER_TO_CODE = {
    'KIRK/DAVID': 'K/D',
    'STEVE L.': 'STL',
    'JOHN': 'JOH',
    'KEVIN': 'KEV',
    'DANNY/JOEY': 'D/J',
    'GREG/GRIFFIN': 'G/G',
    'ERIC/JEFF': 'E/J',
    'ANDREW': 'AND',
    'WES/BILL': 'W/B',
    'KEMP/A/M': 'KAM',
    'JARRETT/MATT': 'J/M',
    'ADAM': 'ADA',
}


def parse_player_name(cell_value: str) -> tuple[str, str]:
    """Parse 'Player Name (TEAM)' into (name, team)."""
    if not cell_value:
        return "", ""
    match = re.match(r'^(.+?)\s*\(([A-Za-z]{2,3})\)$', str(cell_value).strip())
    if match:
        return match.group(1).strip(), match.group(2).upper()
    return str(cell_value).strip(), ""


def extract_team_name(header_value: str) -> tuple[str, str]:
    """Extract team name from header like 'KIRK/DAVID (4)'."""
    if not header_value:
        return "", ""
    match = re.match(r'^(.+?)\s*\((\d+)\)$', str(header_value).strip())
    if match:
        return match.group(1).strip(), match.group(2)
    return str(header_value).strip(), ""


def find_position_rows(ws, start_row: int, end_row: int) -> dict[str, list[int]]:
    """Find rows for each position within a range."""
    position_rows = {}
    current_position = None
    
    for row in range(start_row, min(end_row + 1, ws.max_row + 1)):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value:
            cell_str = str(cell_value).strip().upper()
            if cell_str in POSITIONS:
                current_position = cell_str
                if current_position not in position_rows:
                    position_rows[current_position] = []
                position_rows[current_position].append(row)
        elif current_position:
            has_content = any(
                ws.cell(row=row, column=c).value 
                for c in TEAM_COLUMNS
            )
            if has_content:
                position_rows[current_position].append(row)
    
    return position_rows


def migrate_excel_to_json(excel_path: str, output_dir: str):
    """Convert OPFL Excel data to JSON files."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    # Get all week sheets (OPFL uses W1, W2, ... format)
    week_sheets = []
    for name in wb.sheetnames:
        match = re.match(r'^W(\d+)$', name)
        if match:
            week_sheets.append((int(match.group(1)), name))
    week_sheets.sort(key=lambda x: x[0])
    
    if not week_sheets:
        print("No week sheets found!")
        return
    
    # Use the most recent week for roster extraction
    latest_week_num, latest_sheet = week_sheets[-1]
    ws = wb[latest_sheet]
    
    # Build teams.json and rosters.json
    teams = []
    rosters = {}
    
    # Process both blocks
    blocks = [
        (1, 1, 38),   # Block 1
        (39, 39, 80), # Block 2
    ]
    
    for header_row, start_row, end_row in blocks:
        position_rows = find_position_rows(ws, start_row, end_row)
        
        for col in TEAM_COLUMNS:
            header = ws.cell(row=header_row, column=col).value
            if not header:
                continue
            
            team_name, standing = extract_team_name(header)
            abbrev = OWNER_TO_CODE.get(team_name, team_name[:3].upper())
            
            # Skip if we've already processed this team
            if abbrev in rosters:
                continue
            
            owner_key = team_name.lower().replace(' ', '_').replace('/', '_').replace('.', '')
            owner_key = re.sub(r'[^a-z0-9_]', '', owner_key)
            
            teams.append({
                "abbrev": abbrev,
                "name": team_name,
                "owner": team_name,
                "owner_key": owner_key
            })
            
            # Get roster
            roster = []
            star_col = col - 1
            
            for position, rows in position_rows.items():
                for row in rows:
                    player_cell = ws.cell(row=row, column=col)
                    if player_cell.value:
                        player_name, nfl_team = parse_player_name(str(player_cell.value))
                        if player_name:
                            roster.append({
                                "name": player_name,
                                "nfl_team": nfl_team,
                                "position": position
                            })
            
            rosters[abbrev] = roster
    
    # Save teams.json
    teams_path = output_path / "teams.json"
    with open(teams_path, 'w') as f:
        json.dump({"teams": teams}, f, indent=2)
    print(f"Saved {teams_path}")
    
    # Save rosters.json
    rosters_path = output_path / "rosters.json"
    with open(rosters_path, 'w') as f:
        json.dump(rosters, f, indent=2)
    print(f"Saved {rosters_path}")
    
    # Export lineups for each week
    lineups_dir = output_path / "lineups" / "2025"
    lineups_dir.mkdir(parents=True, exist_ok=True)
    
    for week_num, sheet_name in week_sheets:
        ws = wb[sheet_name]
        week_lineups = {}
        
        for header_row, start_row, end_row in blocks:
            position_rows = find_position_rows(ws, start_row, end_row)
            
            for col in TEAM_COLUMNS:
                header = ws.cell(row=header_row, column=col).value
                if not header:
                    continue
                
                team_name, _ = extract_team_name(header)
                abbrev = OWNER_TO_CODE.get(team_name, team_name[:3].upper())
                
                if abbrev in week_lineups:
                    continue
                
                starters = {pos: [] for pos in POSITIONS}
                star_col = col - 1
                
                for position, rows in position_rows.items():
                    for row in rows:
                        player_cell = ws.cell(row=row, column=col)
                        star_cell = ws.cell(row=row, column=star_col)
                        
                        if player_cell.value:
                            player_name, _ = parse_player_name(str(player_cell.value))
                            is_starter = star_cell.value == '*'
                            if is_starter and player_name:
                                starters[position].append(player_name)
                
                week_lineups[abbrev] = starters
        
        # Save week lineup
        week_path = lineups_dir / f"week_{week_num}.json"
        with open(week_path, 'w') as f:
            json.dump({
                "week": week_num,
                "lineups": week_lineups
            }, f, indent=2)
        print(f"Saved {week_path}")
    
    print("\nMigration complete!")
    print(f"  Teams: {len(teams)}")
    print(f"  Weeks: {len(week_sheets)}")


if __name__ == "__main__":
    import sys
    
    # Default to OPFL Excel file
    excel_files = [
        "Griff OPFL Scoring 2025.xlsx",
        "OPFL Scoring 2025.xlsx",
    ]
    
    excel_path = sys.argv[1] if len(sys.argv) > 1 else None
    if not excel_path:
        for fname in excel_files:
            if Path(fname).exists():
                excel_path = fname
                break
    
    if not excel_path:
        print("Error: No OPFL Excel file found!")
        sys.exit(1)
    
    output_dir = sys.argv[2] if len(sys.argv) > 2 else "data"
    
    migrate_excel_to_json(excel_path, output_dir)
