"""Excel roster parsing utilities for OPFL format."""

import re
from typing import List, Tuple, Optional

import openpyxl

from .models import FantasyTeam
from .constants import DEFENSE_NAME_TO_ABBREV, TEAM_ABBREV_NORMALIZE


def parse_player_name(cell_value: str) -> Tuple[str, str]:
    """
    Parse player name from Excel format "Player Name (TEAM)" to (name, team_abbrev).
    
    Examples:
        "Patrick Mahomes II (KC)" -> ("Patrick Mahomes II", "KC")
        "Caleb Williams (Chi)" -> ("Caleb Williams", "CHI")
        "Baltimore" -> ("Baltimore", "BAL")  # Defense
    """
    if not cell_value:
        return "", ""
    
    cell_value = str(cell_value).strip()
    
    # Check if it's a defense (just team name)
    if cell_value in DEFENSE_NAME_TO_ABBREV:
        return cell_value, DEFENSE_NAME_TO_ABBREV[cell_value]
    
    # Try to match "Player Name (TEAM)" pattern
    match = re.match(r'^(.+?)\s*\(([A-Za-z]{2,3})\)$', cell_value)
    if match:
        name = match.group(1).strip()
        team = match.group(2).upper()
        # Normalize team abbreviation
        team = TEAM_ABBREV_NORMALIZE.get(team, team)
        return name, team
    
    return cell_value, ""


# Valid team columns in OPFL Excel format (columns D, G, J, M, P, S)
# Skip columns V (22) and Y (25) which are duplicate starter-only views
VALID_TEAM_COLUMNS = [4, 7, 10, 13, 16, 19]


def find_team_columns(ws, header_row: int = 1) -> List[Tuple[int, str, int]]:
    """
    Find fantasy team columns in the worksheet.
    
    Only checks the known valid columns (4, 7, 10, 13, 16, 19) to avoid
    picking up duplicate columns like V (22) and Y (25).
    
    Returns:
        List of (column_index, team_name, header_row) tuples
    """
    teams = []
    
    # Only check the valid team columns
    for col in VALID_TEAM_COLUMNS:
        if col <= ws.max_column:
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                cell_str = str(cell_value).strip()
                # Team names in OPFL look like "KIRK/DAVID (11)" or "STEVE L. (43)"
                if re.match(r'^[A-Z\s/\.]+\s*\(\d+\)$', cell_str, re.IGNORECASE):
                    teams.append((col, cell_str, header_row))
    
    return teams


def find_position_rows(ws, start_row: int = 1, end_row: int = 100) -> dict:
    """
    Find the row ranges for each position in the worksheet within a specific range.
    
    Args:
        ws: Worksheet
        start_row: First row to search
        end_row: Last row to search
        
    Returns:
        Dict mapping position -> list of row numbers containing players
    """
    position_rows = {}
    current_position = None
    
    for row in range(start_row, min(end_row + 1, ws.max_row + 1)):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value:
            cell_str = str(cell_value).strip().upper()
            if cell_str in ['QB', 'RB', 'WR', 'TE', 'K', 'DF', 'HC']:
                current_position = cell_str
                if current_position not in position_rows:
                    position_rows[current_position] = []
                # This row contains position players
                position_rows[current_position].append(row)
        elif current_position:
            # Empty column A but might still have players for current position
            # Check if there's actual content in the row
            has_content = any(
                ws.cell(row=row, column=c).value 
                for c in range(2, min(25, ws.max_column + 1))
            )
            if has_content:
                position_rows[current_position].append(row)
    
    return position_rows


def parse_roster_from_excel(filepath: str, sheet_name: str = "W1") -> List[FantasyTeam]:
    """
    Parse fantasy team rosters from OPFL Excel file.
    
    The OPFL Excel format:
        - Teams are arranged horizontally, each occupying 3 columns: Points | Star | Player
        - Two blocks of 6 teams: rows 1-38 and rows 39+
        - Positions are in column A (QB, RB, WR, TE, K, DF, HC)
        - A star (*) in the star column indicates a starter
        - Player names are in format "Name (Team)" or just "TeamName" for defenses
    
    Args:
        filepath: Path to the Excel file
        sheet_name: Name of the sheet to read (e.g., "W1", "W12")
        
    Returns:
        List of FantasyTeam objects
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    
    teams = []
    seen_team_names = set()  # Track teams we've already added
    
    # Find teams in first block (row 1)
    team_columns_block1 = find_team_columns(ws, header_row=1)
    
    # If no teams found via pattern, try the valid columns directly
    if not team_columns_block1:
        for col in VALID_TEAM_COLUMNS:
            if col <= ws.max_column:
                header = ws.cell(row=1, column=col).value
                if header:
                    team_columns_block1.append((col, str(header).strip(), 1))
    
    # Find teams in second block (row 39)
    team_columns_block2 = find_team_columns(ws, header_row=39)
    
    # If no teams found via pattern, try the valid columns directly
    if not team_columns_block2:
        for col in VALID_TEAM_COLUMNS:
            if col <= ws.max_column:
                header = ws.cell(row=39, column=col).value
                if header and str(header).strip():
                    header_str = str(header).strip()
                    if re.match(r'^[A-Z\s/\.]+\s*\(\d+\)$', header_str, re.IGNORECASE):
                        team_columns_block2.append((col, header_str, 39))
    
    # Pre-compute position rows for each block
    position_rows_block1 = find_position_rows(ws, start_row=1, end_row=38)
    position_rows_block2 = find_position_rows(ws, start_row=39, end_row=80)
    
    # Parse teams from block 1
    for player_col, team_name_raw, header_row in team_columns_block1:
        match = re.match(r'^(.+?)\s*\(\d+\)$', team_name_raw)
        team_name = match.group(1).strip() if match else team_name_raw
        
        if team_name in seen_team_names:
            continue
        seen_team_names.add(team_name)
        
        star_col = player_col - 1
        
        team = FantasyTeam(
            name=team_name,
            owner=team_name,
            abbreviation="",
            column_index=player_col,
            players={},
        )
        
        # Use position rows from block 1
        for position, rows in position_rows_block1.items():
            team.players[position] = []
            
            for row in rows:
                player_cell = ws.cell(row=row, column=player_col)
                star_cell = ws.cell(row=row, column=star_col)
                
                if player_cell.value:
                    player_name, nfl_team = parse_player_name(str(player_cell.value))
                    
                    if player_name:
                        is_started = star_cell.value == '*'
                        team.players[position].append((player_name, nfl_team, is_started))
        
        teams.append(team)
    
    # Parse teams from block 2
    for player_col, team_name_raw, header_row in team_columns_block2:
        match = re.match(r'^(.+?)\s*\(\d+\)$', team_name_raw)
        team_name = match.group(1).strip() if match else team_name_raw
        
        if team_name in seen_team_names:
            continue
        seen_team_names.add(team_name)
        
        star_col = player_col - 1
        
        team = FantasyTeam(
            name=team_name,
            owner=team_name,
            abbreviation="",
            column_index=player_col,
            players={},
        )
        
        # Use position rows from block 2
        for position, rows in position_rows_block2.items():
            team.players[position] = []
            
            for row in rows:
                player_cell = ws.cell(row=row, column=player_col)
                star_cell = ws.cell(row=row, column=star_col)
                
                if player_cell.value:
                    player_name, nfl_team = parse_player_name(str(player_cell.value))
                    
                    if player_name:
                        is_started = star_cell.value == '*'
                        team.players[position].append((player_name, nfl_team, is_started))
        
        teams.append(team)
    
    wb.close()
    return teams


def parse_roster_from_rosters_sheet(filepath: str) -> List[FantasyTeam]:
    """
    Parse fantasy team rosters from the 'Rosters' sheet in OPFL Excel file.
    
    Args:
        filepath: Path to the Excel file
        
    Returns:
        List of FantasyTeam objects
    """
    return parse_roster_from_excel(filepath, sheet_name='Rosters')


def update_excel_scores(
    excel_path: str,
    sheet_name: str,
    teams: List[FantasyTeam],
    results: dict,
):
    """
    Update the Excel file with calculated scores.
    
    Args:
        excel_path: Path to the Excel file
        sheet_name: Sheet to update
        teams: List of FantasyTeam objects
        results: Dict mapping team name to (total_score, position_scores)
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]
    
    for team in teams:
        if team.name not in results:
            continue
        
        total, scores = results[team.name]
        player_col = team.column_index
        points_col = player_col - 2
        
        for position, player_list in team.players.items():
            if position not in scores:
                continue
            
            for player_name, nfl_team, is_started in player_list:
                if not is_started:
                    continue
                
                player_score = None
                for ps in scores[position]:
                    if ps.name == player_name:
                        player_score = ps
                        break
                
                if player_score is None:
                    continue
                
                # Find the row for this player
                # Search appropriate range based on team's column location
                for row in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=player_col)
                    if cell.value:
                        parsed_name, _ = parse_player_name(str(cell.value))
                        if parsed_name == player_name:
                            score_cell = ws.cell(row=row, column=points_col)
                            score_cell.value = player_score.total_points
                            break
    
    wb.save(excel_path)
    print(f"\nScores saved to {excel_path}")
