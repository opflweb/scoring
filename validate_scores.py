#!/usr/bin/env python3
"""
Score Validation Script for OPFL

Compares manually entered scores in Excel against calculated scores from nflreadpy.
Outputs a list of discrepancies for review.

Usage:
    python validate_scores.py --sheet "W12" --season 2025 --week 12
    python validate_scores.py --all  # Validate all weeks
"""

import argparse
import re
import sys
from typing import List, Tuple, TextIO

import openpyxl

from opfl import OPFLScorer, parse_roster_from_excel
from opfl.excel_parser import parse_player_name, find_position_rows


class OutputWriter:
    """Write output to both console and file."""
    
    def __init__(self, filepath: str = None):
        self.filepath = filepath
        self.file: TextIO = None
        self.lines: List[str] = []
    
    def __enter__(self):
        if self.filepath:
            self.file = open(self.filepath, 'w', encoding='utf-8')
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.file:
            self.file.close()
    
    def write(self, text: str = ""):
        """Write a line to both console and file."""
        print(text)
        if self.file:
            # Remove emoji for file output (they may not render well in text files)
            clean_text = text.replace('✓', '[OK]').replace('✗', '[X]').replace('⚠', '[!]')
            self.file.write(clean_text + '\n')


def get_excel_scores(filepath: str, sheet_name: str) -> dict:
    """
    Extract manually entered scores from OPFL Excel.
    
    Returns:
        Dict mapping (team_name, position, player_name) -> score
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    
    scores = {}
    seen_keys = set()  # Track unique player entries
    
    # Valid team columns only (skip V=22 and Y=25 which are duplicates)
    valid_columns = [4, 7, 10, 13, 16, 19]
    
    # Find team columns in block 1 (row 1)
    team_columns_block1 = []
    for col in valid_columns:
        if col <= ws.max_column:
            header = ws.cell(row=1, column=col).value
            if header:
                team_columns_block1.append((col, str(header).strip()))
    
    # Find team columns in block 2 (row 39)
    team_columns_block2 = []
    for col in valid_columns:
        if col <= ws.max_column:
            header = ws.cell(row=39, column=col).value
            if header and str(header).strip():
                match = re.match(r'^[A-Z\s/\.]+\s*\(\d+\)$', str(header).strip(), re.IGNORECASE)
                if match:
                    team_columns_block2.append((col, str(header).strip()))
    
    # Find position rows for each block
    position_rows_block1 = find_position_rows(ws, start_row=1, end_row=38)
    position_rows_block2 = find_position_rows(ws, start_row=39, end_row=80)
    
    # Process block 1 teams
    for player_col, team_name_raw in team_columns_block1:
        match = re.match(r'^(.+?)\s*\(\d+\)$', team_name_raw)
        team_name = match.group(1).strip() if match else team_name_raw
        
        star_col = player_col - 1
        points_col = player_col - 2
        
        for position, rows in position_rows_block1.items():
            for row in rows:
                player_cell = ws.cell(row=row, column=player_col)
                star_cell = ws.cell(row=row, column=star_col)
                score_cell = ws.cell(row=row, column=points_col)
                
                if player_cell.value:
                    is_started = star_cell.value == '*'
                    if not is_started:
                        continue
                    
                    player_name, nfl_team = parse_player_name(str(player_cell.value))
                    excel_score = score_cell.value
                    
                    if player_name and excel_score is not None:
                        try:
                            key = (team_name, position, player_name, nfl_team)
                            if key not in seen_keys:
                                scores[key] = float(excel_score)
                                seen_keys.add(key)
                        except (ValueError, TypeError):
                            pass
    
    # Process block 2 teams
    for player_col, team_name_raw in team_columns_block2:
        match = re.match(r'^(.+?)\s*\(\d+\)$', team_name_raw)
        team_name = match.group(1).strip() if match else team_name_raw
        
        star_col = player_col - 1
        points_col = player_col - 2
        
        for position, rows in position_rows_block2.items():
            for row in rows:
                player_cell = ws.cell(row=row, column=player_col)
                star_cell = ws.cell(row=row, column=star_col)
                score_cell = ws.cell(row=row, column=points_col)
                
                if player_cell.value:
                    is_started = star_cell.value == '*'
                    if not is_started:
                        continue
                    
                    player_name, nfl_team = parse_player_name(str(player_cell.value))
                    excel_score = score_cell.value
                    
                    if player_name and excel_score is not None:
                        try:
                            key = (team_name, position, player_name, nfl_team)
                            if key not in seen_keys:
                                scores[key] = float(excel_score)
                                seen_keys.add(key)
                        except (ValueError, TypeError):
                            pass
    
    wb.close()
    return scores


def validate_week(
    excel_path: str,
    sheet_name: str,
    season: int,
    week: int,
    tolerance: float = 0.0,
    output: OutputWriter = None,
) -> Tuple[List[dict], int]:
    """
    Validate scores for a single week.
    
    Args:
        excel_path: Path to Excel file
        sheet_name: Sheet name (e.g., "W12")
        season: NFL season year
        week: Week number
        tolerance: Allow differences up to this amount (default 0 = exact match)
        output: OutputWriter for logging
    
    Returns:
        Tuple of (discrepancies list, total players checked)
    """
    discrepancies = []
    
    # Get Excel scores
    excel_scores = get_excel_scores(excel_path, sheet_name)
    
    if not excel_scores:
        if output:
            output.write(f"  No scored players found in {sheet_name}")
        return discrepancies, 0
    
    total_checked = len(excel_scores)
    
    # Calculate scores
    scorer = OPFLScorer(season, week)
    
    for (team_name, position, player_name, nfl_team), excel_score in excel_scores.items():
        calculated = scorer.score_player(player_name, nfl_team, position)
        
        if not calculated.found_in_stats:
            # Player not found - could be bye week or data issue
            discrepancies.append({
                'week': week,
                'sheet': sheet_name,
                'team': team_name,
                'position': position,
                'player': player_name,
                'nfl_team': nfl_team,
                'excel_score': excel_score,
                'calculated_score': None,
                'difference': None,
                'reason': 'Player not found in stats',
                'breakdown': {},
            })
            continue
        
        diff = excel_score - calculated.total_points
        
        if abs(diff) > tolerance:
            discrepancies.append({
                'week': week,
                'sheet': sheet_name,
                'team': team_name,
                'position': position,
                'player': player_name,
                'nfl_team': nfl_team,
                'excel_score': excel_score,
                'calculated_score': calculated.total_points,
                'difference': diff,
                'reason': 'Score mismatch',
                'breakdown': calculated.breakdown,
                'matched_name': calculated.matched_name,
            })
    
    return discrepancies, total_checked


def get_available_weeks(excel_path: str) -> List[Tuple[str, int]]:
    """Get list of (sheet_name, week_number) for all week sheets."""
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    weeks = []
    
    for sheet_name in wb.sheetnames:
        # Match "W1", "W2", etc.
        match = re.match(r'^W(\d+)$', sheet_name)
        if match:
            week_num = int(match.group(1))
            weeks.append((sheet_name, week_num))
    
    wb.close()
    return sorted(weeks, key=lambda x: x[1])


def print_discrepancies(discrepancies: List[dict], total_checked: int, output: OutputWriter, verbose: bool = True):
    """Pretty print discrepancies with stats."""
    matched = total_checked - len(discrepancies)
    pct = (matched / total_checked * 100) if total_checked > 0 else 0
    
    output.write(f"\n  Checked {total_checked} players: {matched} matched ({pct:.1f}%)")
    
    if not discrepancies:
        output.write("  ✓ All scores match!")
        return
    
    # Group by type
    mismatches = [d for d in discrepancies if d['reason'] == 'Score mismatch']
    not_found = [d for d in discrepancies if d['reason'] == 'Player not found in stats']
    
    if mismatches:
        output.write(f"\n  ⚠ Score Mismatches ({len(mismatches)}):")
        output.write("  " + "-" * 80)
        output.write(f"  {'Pos':<4} {'Player':<30} {'Matched':<20} {'Excel':>7} {'Calc':>7} {'Diff':>7}")
        output.write("  " + "-" * 80)
        
        for d in mismatches:
            player_str = f"{d['player']} ({d['nfl_team']})"[:30]
            matched_str = d.get('matched_name', '')[:20] if d.get('matched_name') != d['player'] else ''
            output.write(f"  {d['position']:<4} {player_str:<30} {matched_str:<20} {d['excel_score']:>7.1f} {d['calculated_score']:>7.1f} {d['difference']:>+7.1f}")
            if verbose and d['breakdown']:
                breakdown_str = ", ".join(f"{k}: {v}" for k, v in d['breakdown'].items() if k != 'floor_applied')
                output.write(f"        └─ {breakdown_str}")
        output.write("")
    
    if not_found:
        output.write(f"\n  ✗ Players Not Found in Stats ({len(not_found)}):")
        output.write("  " + "-" * 70)
        output.write(f"  {'Pos':<4} {'Player':<30} {'Excel':>7} {'Note':<20}")
        output.write("  " + "-" * 70)
        for d in not_found:
            player_str = f"{d['player']} ({d['nfl_team']})"[:30]
            output.write(f"  {d['position']:<4} {player_str:<30} {d['excel_score']:>7.1f} (bye/injured/no stats)")
        output.write("")


def main():
    parser = argparse.ArgumentParser(description="Validate OPFL scores against nflreadpy data")
    parser.add_argument(
        "--excel", "-e",
        default="OPFL Scoring 2025.xlsx",
        help="Path to the Excel file",
    )
    parser.add_argument(
        "--sheet", "-s",
        help="Sheet name to validate (e.g., 'W12')",
    )
    parser.add_argument(
        "--season", "-y",
        type=int,
        default=2025,
        help="NFL season year",
    )
    parser.add_argument(
        "--week", "-w",
        type=int,
        help="Week number (required if --sheet is specified)",
    )
    parser.add_argument(
        "--all", "-a",
        action="store_true",
        help="Validate all weeks",
    )
    parser.add_argument(
        "--tolerance", "-t",
        type=float,
        default=0.0,
        help="Allow score differences up to this amount",
    )
    parser.add_argument(
        "--summary", 
        action="store_true",
        help="Only show summary counts, not individual discrepancies",
    )
    parser.add_argument(
        "--output", "-o",
        default="scoring_validation.txt",
        help="Output file for validation results (default: scoring_validation.txt)",
    )
    parser.add_argument(
        "--no-file",
        action="store_true",
        help="Don't write to output file, only print to console",
    )
    
    args = parser.parse_args()
    
    output_file = None if args.no_file else args.output
    
    with OutputWriter(output_file) as output:
        if args.all:
            # Validate all weeks
            weeks = get_available_weeks(args.excel)
            all_discrepancies = []
            total_players = 0
            
            output.write(f"Validating {len(weeks)} weeks from {args.excel}")
            output.write("=" * 60)
            
            for sheet_name, week_num in weeks:
                output.write(f"\n{sheet_name} (Week {week_num}):")
                discrepancies, checked = validate_week(
                    args.excel, sheet_name, args.season, week_num, args.tolerance, output
                )
                all_discrepancies.extend(discrepancies)
                total_players += checked
                
                if args.summary:
                    mismatches = sum(1 for d in discrepancies if d['reason'] == 'Score mismatch')
                    not_found = sum(1 for d in discrepancies if d['reason'] == 'Player not found in stats')
                    matched = checked - len(discrepancies)
                    pct = (matched / checked * 100) if checked > 0 else 0
                    output.write(f"  {checked} players: {matched} matched ({pct:.1f}%), {mismatches} mismatches, {not_found} not found")
                else:
                    print_discrepancies(discrepancies, checked, output, verbose=True)
            
            # Final summary
            output.write("\n" + "=" * 60)
            output.write("SUMMARY")
            output.write("=" * 60)
            total_mismatches = sum(1 for d in all_discrepancies if d['reason'] == 'Score mismatch')
            total_not_found = sum(1 for d in all_discrepancies if d['reason'] == 'Player not found in stats')
            total_matched = total_players - len(all_discrepancies)
            match_pct = (total_matched / total_players * 100) if total_players > 0 else 0
            output.write(f"Total players checked: {total_players}")
            output.write(f"Total matched: {total_matched} ({match_pct:.1f}%)")
            output.write(f"Total mismatches: {total_mismatches}")
            output.write(f"Total not found: {total_not_found}")
            
            if output_file:
                output.write(f"\nResults written to: {output_file}")
            
        elif args.sheet:
            # Validate single week
            if not args.week:
                # Try to extract week from sheet name
                match = re.match(r'^W(\d+)$', args.sheet)
                if match:
                    args.week = int(match.group(1))
                else:
                    parser.error("--week is required when --sheet is specified")
            
            output.write(f"Validating {args.sheet} (Week {args.week}, Season {args.season})")
            output.write("=" * 60)
            
            discrepancies, checked = validate_week(
                args.excel, args.sheet, args.season, args.week, args.tolerance, output
            )
            print_discrepancies(discrepancies, checked, output, verbose=not args.summary)
            
            if output_file:
                output.write(f"\nResults written to: {output_file}")
            
        else:
            parser.error("Either --sheet or --all is required")


if __name__ == "__main__":
    main()
